import openpyxl
import csv

class FileLoader:
    '''
    FileをロードするためのClass
    '''
    def __init__(self,filepath):
        self.filepath = filepath
    
    def load(self):
        pass
    
    def get(self):
        _val = ""
        with open(self.filepath, mode="r",encoding="utf_8") as text:
            _val = text.read()
        return _val

class CSVLoader(FileLoader):
    '''csvをロードするための'''

    def __init__(self, filepath):
        super().__init__(filepath)
    
    def fetch_data(self):
        with open(self.filepath, 'r', newline="", encoding='shift-jis') as csvfile:
            csv_reader = csv.reader(csvfile)
            header = next(csv_reader)
            data = []
            for row in csv_reader:
                data.append(row)
            
            return data


class ExcelLoader(FileLoader):
    '''
    ExcelをLoadするための
    '''
    def __init__(self,filepath):
        super().__init__(filepath)
        self.wb = openpyxl.load_workbook(filename=filepath, read_only=False)
        self.ws = self.wb.active
        
    def load(self):
        pass
    
    def fetch_data(self, row, col)->str:
        cell_value = str(self.ws.cell(row=row, column=col).value)
        return cell_value
    

    def fetch_selected_data(self, header=False)->list[str]:
        '''
        
        '''
        if header == True:
            start_row = 1
        else:
            start_row = 2
        
        data_list = ""
        for row in self.ws.iter_rows(min_row=start_row, values_only = True):
            data_list = row
        return data_list
    
    def fetch_last_row(self) ->int:
        ''' A列を基準に最終行を取得 '''
        max_row = self.ws.max_row
        last_row = 1
        for row in range(1, max_row + 1):
            cell_value = str(self.ws.cell(row=row, column=1).value)
            if not cell_value:
                break
            last_row = row
        return last_row
    
    def fetch_last_column(self) ->int:
        '''最終列のint型を取得'''
        return self.ws.max_column
    
    def fetch_data_row(self, row)->list[str]:
        return [cell.value for cell in self.ws[row]]
    
    def check_data(self)->None:
        '''
        シート内の全データを出力する関数。
        '''
        last_row = self.fetch_last_row()
        for row in range(2, last_row + 1):
            print(self.fetch_data(row,1))
                  
#test
# loader = ExcelLoader("02sample.xlsx")
# last_row = loader.fetch_last_row()
# loader.check_data()
